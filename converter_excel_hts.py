import openpyxl

def convert_seconds_to_time(seconds):
    """Converte segundos para o formato HH:MM:SS."""
    hours = seconds // 3600
    minutes = (seconds % 3600) // 60
    remaining_seconds = seconds % 60
    return f"{int(hours):02}:{int(minutes):02}:{int(remaining_seconds):02}"

def process_excel(file_path, input_start_column, input_end_column, sheet_name="Sheet1"):
    """Lê valores em segundos de várias colunas e escreve o formato HH:MM:SS em outras células."""
    # Carregar o arquivo Excel
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]

    # Iterar pelas colunas de E até P (ou o intervalo de colunas desejado)
    for row in range(2, sheet.max_row + 1):  # Ignorar o cabeçalho
        for col in range(openpyxl.utils.column_index_from_string(input_start_column),
                         openpyxl.utils.column_index_from_string(input_end_column) + 1):
            cell_value = sheet.cell(row=row, column=col).value
            
            # Verificar se a célula contém um número (pode ser um valor ou fórmula)
            if isinstance(cell_value, (int, float)):
                formatted_time = convert_seconds_to_time(cell_value)
                sheet.cell(row=row, column=col).value = formatted_time
            elif isinstance(cell_value, str) and cell_value.isdigit():  # Verifica se é um número em formato de string
                sheet.cell(row=row, column=col).value = convert_seconds_to_time(int(cell_value))

    # Salvar o arquivo Excel
    wb.save(file_path)
    print(f"Arquivo processado e salvo em: {file_path}")

# Exemplo de uso
file_path = "tempo.xlsx"  # Caminho para o arquivo Excel
input_start_column = "E"  # Primeira coluna (E)
input_end_column = "P"    # Última coluna (P)
process_excel(file_path, input_start_column, input_end_column)
